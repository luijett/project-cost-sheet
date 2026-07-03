"""
项目费用结算 v2 — iOS 18 深色备忘录 + PDF 导入
================================================
"""

import sys, logging
from collections import deque
from decimal import Decimal
from datetime import datetime
from pathlib import Path

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger("CostSheet")

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QPushButton, QLineEdit, QListWidget, QListWidgetItem,
    QScrollArea, QFrame, QSplitter, QMessageBox, QFileDialog,
    QDoubleSpinBox, QDialog, QComboBox, QDialogButtonBox,
    QCheckBox, QStackedWidget, QMenu, QTableWidget, QTableWidgetItem, QHeaderView, QTextEdit,
)
from PySide6.QtCore import Qt, Signal, QTimer, QRect, QPointF, QRectF, QSize, QPoint
from PySide6.QtGui import (QFont, QColor, QPalette, QPainter, QPen, QBrush,
                             QPaintEvent, QPixmap, QIcon, QLinearGradient, QKeyEvent)

from database import Database
from importers import parse_clipboard_text, import_excel
from models import (
    m, calc_tax, CURRENCIES, TAX_RATES, UNIT_PRESETS,
    merge_categories, LineItem,
)

DB_PATH = Path(__file__).parent / "budget.db"
db = Database(DB_PATH)
db.init()

# ═══════════════════════════════════════════
#  配色 — iOS 18 深色备忘录
# ═══════════════════════════════════════════

C = {
    "bg":           "#000000",
    "sidebar_bg":   "#1C1C1E",
    "text":         "#FFFFFF",
    "secondary":    "#98989D",
    "tertiary":     "#636366",
    "accent":       "#FFD60A",
    "accent_hover": "#FFE045",
    "red":          "#FF453A",
    "orange":       "#FF9F0A",
    "separator":    "#38383A",
    "row_bg":       "#1C1C1E",
    "focus_bg":     "#2C2C2E",
    "section":      "#98989D",
    "toolbar_bg":   "rgba(28,28,30,0.95)",
    "summary_bg":   "#0D0D0D",
}

# ── 全局暗色调色板 ──

DARK_PALETTE = QPalette()
DARK_PALETTE.setColor(QPalette.ColorRole.Window, QColor(C["bg"]))
DARK_PALETTE.setColor(QPalette.ColorRole.WindowText, QColor(C["text"]))
DARK_PALETTE.setColor(QPalette.ColorRole.Base, QColor(C["row_bg"]))
DARK_PALETTE.setColor(QPalette.ColorRole.AlternateBase, QColor("#2C2C2E"))
DARK_PALETTE.setColor(QPalette.ColorRole.Text, QColor(C["text"]))
DARK_PALETTE.setColor(QPalette.ColorRole.Button, QColor("#2C2C2E"))
DARK_PALETTE.setColor(QPalette.ColorRole.ButtonText, QColor(C["text"]))
DARK_PALETTE.setColor(QPalette.ColorRole.Highlight, QColor(C["accent"]))
DARK_PALETTE.setColor(QPalette.ColorRole.HighlightedText, QColor("#000000"))
DARK_PALETTE.setColor(QPalette.ColorRole.PlaceholderText, QColor(C["tertiary"]))
DARK_PALETTE.setColor(QPalette.ColorRole.BrightText, QColor(C["red"]))
DARK_PALETTE.setColor(QPalette.ColorRole.Link, QColor(C["accent"]))
DARK_PALETTE.setColor(QPalette.ColorGroup.Disabled, QPalette.ColorRole.Text, QColor("#636366"))
DARK_PALETTE.setColor(QPalette.ColorGroup.Disabled, QPalette.ColorRole.ButtonText, QColor("#636366"))

STYLE = f"""
* {{ font-family: "Microsoft YaHei", "PingFang SC", "SF Pro Text", sans-serif; }}
QMainWindow {{ background: {C['bg']}; }}
QWidget {{ color: {C['text']}; }}

/* ── 侧边栏 ── */
#sidebar {{
    background: {C['sidebar_bg']};
    border-right: 0.5px solid {C['separator']};
}}
#sidebarTitle {{
    font-size: 28px; font-weight: 800; color: {C['text']};
    padding: 4px 0; letter-spacing: 0.2px;
}}
#projectList {{
    background: transparent; border: none; outline: none; font-size: 15px; font-weight: 600;
}}
#projectList::item {{
    padding: 11px 16px; border-radius: 10px; margin: 1px 8px;
    color: {C['text']}; font-weight: 600;
}}
#projectList::item:selected {{ background: {C['accent']}; color: #000000; font-weight: 700; }}
#projectList::item:hover:!selected {{ background: rgba(255,214,10,0.12); }}

/* ── 新增按钮（圆角胶囊） ── */
#newProjectBtn {{
    background: {C['accent']}; color: #000000; border: none;
    border-radius: 24px; padding: 9px 20px;
    font-size: 15px; font-weight: 700;
}}
#newProjectBtn:hover {{ background: {C['accent_hover']}; }}
#newProjectBtn:pressed {{ background: #E5C008; }}

/* ── 通用按钮 ── */
#primaryBtn {{
    background: {C['accent']}; color: #000000; border: none;
    border-radius: 22px; padding: 9px 22px; font-size: 14px; font-weight: 700;
}}
#primaryBtn:hover {{ background: {C['accent_hover']}; }}
#primaryBtn:pressed {{ background: #E5C008; }}

/* VS Code 风格标签按钮 — 默认透明无框，hover 深灰底 */
QPushButton#tabBtn {{
    background: transparent; color: {C['text']}; border: none;
    border-radius: 8px; padding: 7px 14px; font-size: 13px; font-weight: 600;
}}
QPushButton#tabBtn:hover {{ background: rgba(255,255,255,0.08); }}
QPushButton#tabBtn:pressed {{ background: rgba(255,255,255,0.14); }}

#linkBtn {{
    background: transparent; color: {C['accent']}; border: none;
    font-size: 13px; padding: 6px 4px; text-align: left; font-weight: 600;
}}
#linkBtn:hover {{ color: {C['accent_hover']}; }}
#linkBtn:pressed {{ color: #E5C008; }}

#deleteBtn {{
    background: transparent; color: {C['red']}; border: none;
    font-size: 16px; padding: 4px 8px; border-radius: 12px; font-weight: 700;
}}
#deleteBtn:hover {{ background: rgba(255,69,58,0.12); }}

/* ── 输入框 ── */
QLineEdit {{
    border: none; border-radius: 6px; padding: 7px 10px;
    font-size: 14px; background: {C['row_bg']}; color: {C['text']};
    selection-background-color: {C['accent']}; selection-color: #000000;
}}
QLineEdit:focus {{ background: {C['focus_bg']}; }}

QDoubleSpinBox {{
    border: none; border-radius: 6px; padding: 6px 8px;
    font-size: 14px; background: {C['row_bg']}; color: {C['text']};
}}
QDoubleSpinBox:focus {{ background: {C['focus_bg']}; }}
QDoubleSpinBox QLineEdit {{ color: {C['text']}; }}

QSpinBox {{
    border: none; border-radius: 6px; padding: 6px 8px;
    font-size: 14px; background: {C['row_bg']}; color: {C['text']};
}}
QSpinBox:focus {{ background: {C['focus_bg']}; }}

/* ── 下拉框（无背景，由各实例自行设置） ── */
QComboBox {{
    border: none; border-radius: 6px; padding: 6px 10px;
    font-size: 13px; font-weight: 600; color: {C['text']};
}}
QComboBox QAbstractItemView {{
    background: #1C1C1E; border: 1px solid {C['separator']};
    border-radius: 8px; selection-background-color: {C['accent']};
    selection-color: #000000; outline: none; padding: 4px;
    color: {C['text']}; font-weight: 600;
}}
QComboBox QLineEdit {{ color: {C['text']}; background: transparent; padding: 0px; }}

/* ── 复选框 — 尺寸交给 Qt 原生计算，只设颜色 ── */
QCheckBox {{
    font-size: 17px; font-weight: 600; spacing: 8px; color: {C['text']};
    padding: 6px 0px;
}}
QCheckBox::indicator {{
    border: 2px solid #48484A; border-radius: 4px; background: transparent;
}}
QCheckBox::indicator:checked {{
    background: {C['accent']}; border-color: {C['accent']};
}}
QCheckBox::indicator:hover {{ border-color: {C['accent']}; }}

/* ── 滚动条 ── */
QScrollArea {{ border: none; background: {C['bg']}; }}
QScrollBar:vertical {{ width: 5px; border: none; background: transparent; }}
QScrollBar::handle:vertical {{
    background: rgba(255,255,255,0.18); border-radius: 3px; min-height: 30px;
}}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{ height: 0; }}

/* ── 对话框 ── */
QDialog {{ background: #1C1C1E; border-radius: 14px; }}
QMessageBox {{ background: #1C1C1E; color: {C['text']}; }}
QMessageBox QLabel {{ color: {C['text']}; font-weight: 600; }}

/* ── 汇总栏 ── */
#summaryBar {{
    background: {C['summary_bg']};
    border-top: 0.5px solid {C['separator']};
}}
"""


# ═══════════════════════════════════════════
#  自定义绘制 + 号圆形按钮
# ═══════════════════════════════════════════

class CirclePlusButton(QPushButton):
    def __init__(self, color="#FFD60A", hover_color="#FFE045", press_color="#E5C008", parent=None):
        super().__init__(parent)
        self._color = QColor(color)
        self._hover_color = QColor(hover_color)
        self._press_color = QColor(press_color)
        self._hovered = False
        self._pressed = False
        self.setFixedSize(34, 34)
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        # 去掉默认样式，完全自绘
        self.setStyleSheet("background: transparent; border: none;")

    def enterEvent(self, event):
        self._hovered = True; self.update(); super().enterEvent(event)

    def leaveEvent(self, event):
        self._hovered = False; self.update(); super().leaveEvent(event)

    def mousePressEvent(self, event):
        self._pressed = True; self.update(); super().mousePressEvent(event)

    def mouseReleaseEvent(self, event):
        self._pressed = False; self.update(); super().mouseReleaseEvent(event)

    def paintEvent(self, event: QPaintEvent):
        p = QPainter(self)
        p.setRenderHint(QPainter.RenderHint.Antialiasing)

        # 选色
        if self._pressed:
            bg = self._press_color
        elif self._hovered:
            bg = self._hover_color
        else:
            bg = self._color

        # 画圆
        p.setBrush(QBrush(bg))
        p.setPen(Qt.PenStyle.NoPen)
        p.drawEllipse(2, 2, 30, 30)

        # 画 +
        p.setPen(QPen(QColor("#000000"), 3, Qt.PenStyle.SolidLine, Qt.PenCapStyle.RoundCap))
        cx, cy = 17, 17   # 圆心
        half = 7            # + 的半长
        p.drawLine(cx - half, cy, cx + half, cy)   # 横
        p.drawLine(cx, cy - half, cx, cy + half)   # 竖

        p.end()


# ═══════════════════════════════════════════
#  iOS 风格图标绘制
# ═══════════════════════════════════════════

def _draw_icon(size, color, draw_fn):
    """2x 渲染 → SmoothTransformation 缩到 1x，高清不放大"""
    big = size * 2
    px = QPixmap(big, big)
    px.fill(Qt.GlobalColor.transparent)
    p = QPainter(px)
    p.setRenderHint(QPainter.RenderHint.Antialiasing)
    p.scale(2, 2)
    draw_fn(p, QColor(color), size)
    p.end()
    return px.scaled(size, size, Qt.AspectRatioMode.KeepAspectRatio,
                     Qt.TransformationMode.SmoothTransformation)


def ios_camera_icon(size=22, color="#FFD60A"):
    """平面 — 镜头光圈"""
    def draw(p, c, s):
        cx, cy = s/2, s/2
        # 外环
        p.setPen(QPen(c, 1.2, Qt.PenStyle.SolidLine, Qt.PenCapStyle.RoundCap))
        p.setBrush(Qt.BrushStyle.NoBrush)
        p.drawEllipse(QPointF(cx, cy), s/2 - 2.5, s/2 - 2.5)
        # 内环
        p.setPen(QPen(c, 1.0))
        p.drawEllipse(QPointF(cx, cy), s/6, s/6)
        # 光圈叶片
        p.setPen(QPen(c, 0.8))
        for angle in [45, 135, 225, 315]:
            import math
            rad = math.radians(angle)
            mr = s/2 - 4; ir = s/6 + 0.5
            p.drawLine(QPointF(cx + ir*math.cos(rad), cy - ir*math.sin(rad)),
                       QPointF(cx + mr*math.cos(rad), cy - mr*math.sin(rad)))
    return _draw_icon(size, color, draw)


def ios_video_icon(size=22, color="#FFD60A"):
    """视频 — 16:9 框 + 播放三角"""
    def draw(p, c, s):
        pen = QPen(c, 1.2, Qt.PenStyle.SolidLine, Qt.PenCapStyle.RoundCap, Qt.PenJoinStyle.RoundJoin)
        p.setPen(pen); p.setBrush(Qt.BrushStyle.NoBrush)
        m = 2; w = s - 2*m; h = w * 9/16; y = (s - h)/2
        p.drawRoundedRect(QRectF(m, y, w, h), 4, 4)
        # 播放三角
        tcx, tcy = s/2 + 0.5, s/2; tr = s/8
        pts = [QPointF(tcx - tr*0.55, tcy - tr),
               QPointF(tcx + tr*0.85, tcy),
               QPointF(tcx - tr*0.55, tcy + tr)]
        p.setBrush(QBrush(c)); p.setPen(Qt.PenStyle.NoPen)
        p.drawPolygon(pts)
    return _draw_icon(size, color, draw)


def ios_combo_icon(size=22, color="#FFD60A"):
    """套拍 — 小光圈 + 小播放框并排"""
    def draw(p, c, s):
        pen = QPen(c, 1.2, Qt.PenStyle.SolidLine, Qt.PenCapStyle.RoundCap, Qt.PenJoinStyle.RoundJoin)
        p.setPen(pen); p.setBrush(Qt.BrushStyle.NoBrush)
        half = s/2; gap = 2
        l_cx = half/2 + gap; l_cy = s/2; l_r = half * 0.52
        p.drawEllipse(QPointF(l_cx, l_cy), l_r, l_r)
        p.setPen(QPen(c, 0.9)); p.drawEllipse(QPointF(l_cx, l_cy), l_r/3, l_r/3)
        p.setPen(pen)
        r_x = half + gap; r_w = half * 0.7; r_h = r_w * 9/16; r_y = l_cy - r_h/2
        p.drawRoundedRect(QRectF(r_x, r_y, r_w, r_h), 2.5, 2.5)
        tcx = r_x + r_w/2 + 0.5; tcy = l_cy; tr = r_h/3.5
        pts = [QPointF(tcx - tr*0.5, tcy - tr),
               QPointF(tcx + tr*0.75, tcy),
               QPointF(tcx - tr*0.5, tcy + tr)]
        p.setBrush(QBrush(c)); p.setPen(Qt.PenStyle.NoPen); p.drawPolygon(pts)
    return _draw_icon(size, color, draw)


# ═══════════════════════════════════════════
#  新建项目对话框
# ═══════════════════════════════════════════

class NewProjectDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("新建项目")
        self.setFixedSize(620, 370)
        self.setModal(True)
        self.setStyleSheet(f"""
            QDialog {{ background: #1C1C1E; border-radius: 14px; }}
            QLabel {{ color: {C['text']}; font-weight: 600; }}
            QLineEdit {{
                border: 1.5px solid {C['separator']}; border-radius: 10px;
                padding: 12px 16px; font-size: 16px; font-weight: 600;
                background: #000000; color: {C['text']}; margin: 8px 0;
                selection-background-color: {C['accent']}; selection-color: #000000;
            }}
            QLineEdit:focus {{ border: 2px solid {C['accent']}; padding: 11px 15px; }}
            QCheckBox {{
                font-size: 17px; font-weight: 600; spacing: 8px; color: {C['text']};
                padding: 6px 0px;
            }}
            QCheckBox::indicator {{
                border: 2px solid #48484A; border-radius: 4px; background: transparent;
            }}
            QCheckBox::indicator:checked {{
                background: {C['accent']}; border-color: {C['accent']};
            }}
            QCheckBox::indicator:hover {{ border-color: {C['accent']}; }}
        """)

        lo = QVBoxLayout(self)
        lo.setContentsMargins(28, 28, 28, 24); lo.setSpacing(16)

        t = QLabel("新建项目"); t.setStyleSheet("font-size:22px;font-weight:800;")
        lo.addWidget(t)

        self.name_edit = QLineEdit()
        self.name_edit.setPlaceholderText("项目名称（如：Nike 2024SS）")
        lo.addWidget(self.name_edit)

        tl = QLabel("项目类型（可多选，支持套拍）")
        tl.setStyleSheet(f"font-size:13px;color:{C['secondary']};font-weight:600;")
        lo.addWidget(tl)

        tr = QHBoxLayout(); tr.setSpacing(28)
        self.cb_print = QCheckBox("平面拍摄"); self.cb_print.setChecked(True)
        self.cb_video = QCheckBox("视频拍摄")
        tr.addStretch(); tr.addWidget(self.cb_print); tr.addWidget(self.cb_video); tr.addStretch()
        lo.addLayout(tr)

        br = QHBoxLayout(); br.addStretch()
        cancel = QPushButton("取消")
        cancel.setStyleSheet(
            f"background:transparent;color:{C['secondary']};border:none;font-size:15px;font-weight:600;padding:10px 18px;")
        cancel.clicked.connect(self.reject); br.addWidget(cancel)
        create = QPushButton("创建")
        create.setStyleSheet(f"""
            QPushButton {{ background: {C['accent']}; color: #000000; border: none;
                           border-radius: 20px; padding: 10px 28px; font-size: 15px; font-weight: 700; }}
            QPushButton:hover {{ background: {C['accent_hover']}; }}
            QPushButton:pressed {{ background: #E5C008; }}
        """)
        create.clicked.connect(self._ok); br.addWidget(create)
        lo.addLayout(br)

    def _ok(self):
        if not self.name_edit.text().strip():
            QMessageBox.warning(self, "提示", "请输入项目名称")
            return
        if not self.cb_print.isChecked() and not self.cb_video.isChecked():
            QMessageBox.warning(self, "提示", "至少选择一个项目类型")
            return
        self.accept()

    def get_data(self):
        name = self.name_edit.text().strip()
        types = []
        if self.cb_print.isChecked(): types.append("print")
        if self.cb_video.isChecked(): types.append("video")
        return name, ",".join(types)


# ═══════════════════════════════════════════
#  费用行（非粗体 — 数据区域保持清爽）
# ═══════════════════════════════════════════

class LineItemRow(QFrame):
    changed = Signal()
    remove_requested = Signal(object)

    def __init__(self, item_data=None, currency="CNY"):
        super().__init__()
        self.item_data = item_data or LineItem()
        self.currency = currency; self._busy = False
        self._setup()
        if item_data: self._load()

    def _setup(self):
        row = QHBoxLayout(self); row.setContentsMargins(0, 3, 0, 3); row.setSpacing(8)

        self.desc = QLineEdit()
        self.desc.setPlaceholderText("描述"); self.desc.setMinimumWidth(140)
        self.desc.textChanged.connect(self._on_change)
        row.addWidget(self.desc, 3)

        sym = CURRENCIES.get(self.currency, {}).get("symbol", "¥")
        self.price = QDoubleSpinBox()
        self.price.setRange(0, 99999999); self.price.setDecimals(2)
        self.price.setPrefix(f"{sym} "); self.price.setMinimumWidth(115)
        self.price.valueChanged.connect(self._on_change)
        row.addWidget(self.price)

        self.qty = QDoubleSpinBox()
        self.qty.setRange(0.01, 99999); self.qty.setDecimals(2)
        self.qty.setValue(1); self.qty.setFixedWidth(70)
        self.qty.valueChanged.connect(self._on_change)
        row.addWidget(self.qty)

        self.unit_combo = QComboBox()
        self.unit_combo.setEditable(True)
        self.unit_combo.addItems(UNIT_PRESETS)
        self.unit_combo.setCurrentText(""); self.unit_combo.setFixedWidth(70)
        self.unit_combo.setStyleSheet(f"""
            QComboBox {{ background: transparent; color: {C['text']}; border: none;
                        border-radius: 6px; padding: 6px 10px; font-size: 13px; font-weight: 600; }}
            QComboBox:hover {{ background: rgba(255,255,255,0.08); }}
        """)
        self.unit_combo.currentTextChanged.connect(self._on_change)
        self.unit_combo.lineEdit().setPlaceholderText("单位")
        row.addWidget(self.unit_combo)

        self.total_label = QLabel("¥0")
        self.total_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.total_label.setMinimumWidth(100)
        self.total_label.setStyleSheet(f"font-size:14px;color:{C['text']};")
        row.addWidget(self.total_label)

        db_btn = QPushButton("✕"); db_btn.setObjectName("deleteBtn"); db_btn.setFixedSize(30, 30)
        db_btn.clicked.connect(lambda: self.remove_requested.emit(self))
        row.addWidget(db_btn)

    def _on_change(self):
        if self._busy: return
        self._calc(); self.changed.emit()

    def _calc(self):
        total = m(self.price.value()) * m(self.qty.value())
        sym = CURRENCIES.get(self.currency, {}).get("symbol", "¥")
        self.total_label.setText(f"{sym}{total:,.2f}")
        self.item_data.unit_price = m(self.price.value())
        self.item_data.quantity = m(self.qty.value())
        self.item_data.total = total

    def _load(self):
        self._busy = True
        self.desc.setText(self.item_data.description)
        self.price.setValue(float(self.item_data.unit_price))
        self.qty.setValue(float(self.item_data.quantity))
        idx = self.unit_combo.findText(self.item_data.unit)
        if idx >= 0: self.unit_combo.setCurrentIndex(idx)
        elif self.item_data.unit: self.unit_combo.setCurrentText(self.item_data.unit)
        self._busy = False; self._calc()

    def set_currency(self, currency):
        self.currency = currency
        self.price.setPrefix(f"{CURRENCIES.get(currency, {}).get('symbol', '¥')} ")
        self._calc()

    def get_data(self) -> LineItem:
        self.item_data.description = self.desc.text().strip()
        self.item_data.unit_price = m(self.price.value())
        self.item_data.quantity = m(self.qty.value())
        self.item_data.unit = self.unit_combo.currentText().strip()
        self.item_data.total = self.item_data.unit_price * self.item_data.quantity
        return self.item_data

    def total(self) -> Decimal:
        return self.item_data.total or m(self.price.value()) * m(self.qty.value())


# ═══════════════════════════════════════════
#  分类区块
# ═══════════════════════════════════════════

class CategorySection(QFrame):
    changed = Signal()

    def __init__(self, category_data=None, default_items=None, currency="CNY"):
        super().__init__()
        self.category_data = category_data or {}
        self.currency = currency
        self.item_widgets: list[LineItemRow] = []
        self._collapsed = False
        self._setup()
        if default_items:
            for desc, note in default_items:
                self.add_item(LineItem(description=desc, notes=note))
        if not self.item_widgets: self.add_item()

    def _setup(self):
        self.setStyleSheet("CategorySection { background: transparent; }")
        lo = QVBoxLayout(self); lo.setContentsMargins(0, 0, 0, 0); lo.setSpacing(0)

        hdr = QWidget()
        hdr.setCursor(Qt.CursorShape.PointingHandCursor)
        hdr.mousePressEvent = lambda e: self._toggle()
        hl = QHBoxLayout(hdr); hl.setContentsMargins(32, 20, 24, 6)
        self.toggle_btn = QLabel("▼")
        self.toggle_btn.setStyleSheet(f"font-size:10px;color:{C['secondary']};padding-right:6px;")
        hl.addWidget(self.toggle_btn)
        nm = QLabel(self.category_data.get("name", "").upper())
        nm.setStyleSheet(
            f"font-size:12px;font-weight:800;color:{C['section']};letter-spacing:0.6px;")
        hl.addWidget(nm); hl.addStretch()
        self.subtotal_label = QLabel("¥0")
        self.subtotal_label.setStyleSheet(
            f"font-size:12px;font-weight:700;color:{C['secondary']};")
        hl.addWidget(self.subtotal_label); lo.addWidget(hdr)

        self._sep = QFrame(); self._sep.setFrameShape(QFrame.Shape.HLine)
        self._sep.setStyleSheet(f"background:{C['separator']};max-height:0.5px;margin:0 24px;")
        lo.addWidget(self._sep)

        self.items_widget = QWidget()
        self.items_layout = QVBoxLayout(self.items_widget)
        self.items_layout.setContentsMargins(24, 4, 16, 4); self.items_layout.setSpacing(0)
        lo.addWidget(self.items_widget)

        self._add_btn = QPushButton("+ 添加项目")
        self._add_btn.setObjectName("linkBtn"); self._add_btn.setStyleSheet(
            f"background:transparent;color:{C['accent']};border:none;font-size:13px;font-weight:700;padding:6px 4px;")
        self._add_btn.clicked.connect(lambda: self.add_item()); lo.addWidget(self._add_btn)

        self._spacer = QWidget(); self._spacer.setFixedHeight(12); lo.addWidget(self._spacer)

    def _toggle(self):
        self._collapsed = not self._collapsed
        self.toggle_btn.setText("▶" if self._collapsed else "▼")
        self.items_widget.setVisible(not self._collapsed)
        self._add_btn.setVisible(not self._collapsed)
        self._sep.setVisible(not self._collapsed)
        self._spacer.setVisible(not self._collapsed)

    def add_item(self, item_data=None):
        w = LineItemRow(item_data, self.currency)
        w.changed.connect(self._on_item_change)
        w.remove_requested.connect(self._remove_item)
        self.items_layout.addWidget(w); self.item_widgets.append(w)
        self._update_subtotal(); self.changed.emit()
        w.desc.setFocus()
        return w

    def _remove_item(self, widget):
        if len(self.item_widgets) <= 1:
            widget.desc.clear(); widget.price.setValue(0)
            widget.qty.setValue(1); widget.unit_combo.setCurrentIndex(0)
            widget._calc(); return
        self.items_layout.removeWidget(widget)
        self.item_widgets.remove(widget); widget.deleteLater()
        self._update_subtotal(); self.changed.emit()

    def _on_item_change(self):
        self._update_subtotal(); self.changed.emit()

    def _update_subtotal(self):
        total = sum(w.total() for w in self.item_widgets)
        sym = CURRENCIES.get(self.currency, {}).get("symbol", "¥")
        self.subtotal_label.setText(f"{sym}{total:,.2f}")

    def set_currency(self, currency):
        self.currency = currency
        for w in self.item_widgets: w.set_currency(currency)
        self._update_subtotal()

    def subtotal(self) -> Decimal:
        return sum(w.total() for w in self.item_widgets)

    def get_items(self) -> list[LineItem]:
        return [w.get_data() for w in self.item_widgets]


# ═══════════════════════════════════════════
#  主窗口
# ═══════════════════════════════════════════

class App(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("项目费用结算 v2")
        self.resize(1200, 840)
        screen = self.screen().availableGeometry()
        self.move((screen.width() - self.width()) // 2, (screen.height() - self.height()) // 2)

        self.current_pid = None; self.currency = "CNY"; self.tax_rate = Decimal("0")
        self.sections: list[CategorySection] = []
        self._dirty = False
        self._timer = QTimer(); self._timer.setSingleShot(True)
        self._timer.timeout.connect(self._save)
        self._undo_stack = deque(maxlen=50)
        self._undoing = False
        self._loading = False

        self._setup_ui(); self._load_projects()

    def _setup_ui(self):
        splitter = QSplitter(Qt.Orientation.Horizontal)
        splitter.setHandleWidth(1)

        # ── 侧边栏 ──
        sidebar = QWidget(); sidebar.setObjectName("sidebar"); sidebar.setFixedWidth(260)
        sl = QVBoxLayout(sidebar); sl.setContentsMargins(0, 0, 0, 0); sl.setSpacing(0)

        top = QWidget()
        tl = QHBoxLayout(top); tl.setContentsMargins(20, 24, 16, 14)
        t = QLabel("项目"); t.setObjectName("sidebarTitle")
        tl.addWidget(t); tl.addStretch()
        add = CirclePlusButton(color=C['accent'], hover_color=C['accent_hover'], press_color="#E5C008")
        add.clicked.connect(self._new_project); tl.addWidget(add)
        sl.addWidget(top)

        self.project_list = QListWidget()
        self.project_list.setObjectName("projectList")
        self.project_list.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.project_list.setTextElideMode(Qt.TextElideMode.ElideRight)
        self.project_list.setDragDropMode(self.project_list.DragDropMode.InternalMove)
        self.project_list.setDefaultDropAction(Qt.DropAction.MoveAction)
        self.project_list.model().rowsMoved.connect(self._on_projects_reordered)
        self.project_list.currentRowChanged.connect(self._on_select)
        sl.addWidget(self.project_list, 1)

        bt = QWidget()
        bl = QHBoxLayout(bt); bl.setContentsMargins(16, 8, 16, 14)
        db_btn = QPushButton("DELETE")
        db_btn.setStyleSheet(f"""
            QPushButton {{ background: transparent; color: {C['red']}; border: none;
                          border-radius: 8px; padding: 7px 14px;
                          font-size: 12px; font-weight: 700; letter-spacing: 0.5px; }}
            QPushButton:hover {{ background: rgba(255,69,58,0.10); }}
            QPushButton:pressed {{ background: rgba(255,69,58,0.18); }}
        """)
        db_btn.clicked.connect(self._delete_project)
        bl.addWidget(db_btn); bl.addStretch(); sl.addWidget(bt)
        splitter.addWidget(sidebar)

        # ── 右侧 ──
        right = QWidget()
        right.setStyleSheet(f"background:{C['bg']};")
        rl = QVBoxLayout(right); rl.setContentsMargins(0, 0, 0, 0); rl.setSpacing(0)

        self.stack = QStackedWidget()

        # 空状态
        ep = QWidget()
        el = QVBoxLayout(ep); el.setAlignment(Qt.AlignmentFlag.AlignCenter)
        ic = QLabel("📋"); ic.setStyleSheet("font-size:48px;color:#48484A;")
        ic.setAlignment(Qt.AlignmentFlag.AlignCenter); el.addWidget(ic)
        tx = QLabel("选择左侧项目或点击 新增")
        tx.setStyleSheet(f"font-size:15px;font-weight:600;color:{C['tertiary']};")
        tx.setAlignment(Qt.AlignmentFlag.AlignCenter); el.addWidget(tx)
        self.stack.addWidget(ep)

        # 编辑页
        ed = QWidget()
        ev = QVBoxLayout(ed); ev.setContentsMargins(0, 0, 0, 0); ev.setSpacing(0)

        # ── 工具栏 ──
        tbw = QWidget()
        tbw.setStyleSheet(
            f"background:{C['toolbar_bg']};border-bottom:0.5px solid {C['separator']};")
        tb = QHBoxLayout(tbw); tb.setContentsMargins(24, 12, 20, 12); tb.setSpacing(14)

        self.proj_icon = QLabel()
        self.proj_icon.setFixedSize(26, 26)
        self.proj_icon.setStyleSheet("background:transparent; border:none;")
        self.proj_icon.setPixmap(ios_camera_icon(26, C['accent']))
        tb.addWidget(self.proj_icon)

        self.proj_name_edit = QLineEdit()
        self.proj_name_edit.setStyleSheet(f"""
            font-size: 18px; font-weight: 800; color: {C['text']};
            background: transparent; border: none; border-radius: 6px;
            padding: 4px 8px;
        """)
        self.proj_name_edit.textChanged.connect(self._on_name_edit)
        tb.addWidget(self.proj_name_edit, 1)

        cl = QLabel("币种")
        cl.setStyleSheet(f"font-size:12px;font-weight:700;color:{C['secondary']};background:transparent;border:none;"); tb.addWidget(cl)
        self.cur_combo = QComboBox()
        self.cur_combo.setStyleSheet(f"""
            QComboBox {{ background: transparent; color: {C['text']}; border: none;
                        border-radius: 8px; padding: 6px 10px; font-size: 13px; font-weight: 700; }}
            QComboBox:hover {{ background: rgba(255,255,255,0.08); }}
            QComboBox QAbstractItemView {{
                background: #1C1C1E; border: 1px solid {C['separator']};
                border-radius: 8px; selection-background-color: {C['accent']};
                selection-color: #000000; padding: 4px; color: {C['text']};
            }}
        """)
        for code, info in CURRENCIES.items():
            self.cur_combo.addItem(f"{info['symbol']} {code}", code)
        self.cur_combo.currentIndexChanged.connect(self._on_currency_change)
        tb.addWidget(self.cur_combo)

        accent_btn_style = f"""
            QPushButton {{ background: transparent; color: {C['accent']}; border: none;
                          border-radius: 8px; padding: 7px 14px; font-size: 13px; font-weight: 700; }}
            QPushButton:hover {{ background: rgba(255,214,10,0.10); }}
            QPushButton:pressed {{ background: rgba(255,214,10,0.18); }}
        """
        self.undo_btn = QPushButton("↩ 撤销")
        self.undo_btn.setEnabled(False)
        self.undo_btn.setStyleSheet(f"""
            QPushButton {{ background: transparent; color: {C['secondary']}; border: none;
                          border-radius: 8px; padding: 7px 14px; font-size: 13px; font-weight: 600; }}
            QPushButton:hover {{ background: rgba(255,255,255,0.08); }}
            QPushButton:pressed {{ background: rgba(255,255,255,0.14); }}
            QPushButton:disabled {{ color: #48484A; }}
        """)
        self.undo_btn.clicked.connect(self._undo)
        tb.addWidget(self.undo_btn)

        for label, slot in [("粘贴导入", self._import_paste), ("导入 Excel", self._import_excel),
                            ("导出 Excel", self._export_excel), ("导出 PDF", self._export_pdf)]:
            btn = QPushButton(label)
            btn.setStyleSheet(accent_btn_style)
            btn.clicked.connect(slot)
            tb.addWidget(btn)

        ev.addWidget(tbw)

        # ── 预算内容 ──
        self.scroll = QScrollArea(); self.scroll.setWidgetResizable(True)
        self.scroll_content = QWidget()
        self.scroll_content.setStyleSheet("background:transparent;")
        self.cards_layout = QVBoxLayout(self.scroll_content)
        self.cards_layout.setContentsMargins(0, 12, 0, 12)
        self.cards_layout.setSpacing(0); self.cards_layout.addStretch()
        self.scroll.setWidget(self.scroll_content)
        ev.addWidget(self.scroll, 1)

        # ── 汇总栏 ──
        self.summary_bar = QWidget(); self.summary_bar.setObjectName("summaryBar")
        sb = QVBoxLayout(self.summary_bar)
        sb.setContentsMargins(32, 16, 28, 18); sb.setSpacing(8)

        r1 = QHBoxLayout()
        l1 = QLabel("不含税合计")
        l1.setStyleSheet(f"font-size:13px;font-weight:700;color:{C['secondary']};")
        r1.addWidget(l1); r1.addStretch()
        self.subtotal_label = QLabel("¥0")
        self.subtotal_label.setStyleSheet(f"font-size:15px;font-weight:700;color:{C['text']};")
        r1.addWidget(self.subtotal_label); sb.addLayout(r1)

        r2 = QHBoxLayout()
        l2 = QLabel("税率")
        l2.setStyleSheet(f"font-size:13px;font-weight:700;color:{C['secondary']};")
        r2.addWidget(l2)
        self.tax_combo = QComboBox()
        self.tax_combo.setStyleSheet(f"""
            QComboBox {{ background: transparent; color: {C['text']}; border: none;
                        border-radius: 8px; padding: 6px 10px; font-size: 13px; font-weight: 700; }}
            QComboBox:hover {{ background: rgba(255,255,255,0.08); }}
            QComboBox QAbstractItemView {{
                background: #1C1C1E; border: 1px solid {C['separator']};
                border-radius: 8px; selection-background-color: {C['accent']};
                selection-color: #000000; padding: 4px; color: {C['text']};
            }}
        """)
        for label, rate in TAX_RATES:
            self.tax_combo.addItem(label, str(rate))
        self.tax_combo.currentIndexChanged.connect(self._on_tax_change)
        r2.addWidget(self.tax_combo); r2.addStretch()
        self.tax_label = QLabel("¥0")
        self.tax_label.setStyleSheet(f"font-size:15px;font-weight:700;color:{C['orange']};")
        r2.addWidget(self.tax_label); sb.addLayout(r2)

        sep_line = QFrame(); sep_line.setFrameShape(QFrame.Shape.HLine)
        sep_line.setStyleSheet(f"background:{C['separator']};max-height:0.5px;margin:4px 0;")
        sb.addWidget(sep_line)

        r3 = QHBoxLayout()
        l3 = QLabel("含税总计")
        l3.setStyleSheet(f"font-size:15px;font-weight:800;color:{C['text']};")
        r3.addWidget(l3); r3.addStretch()
        self.grand_label = QLabel("¥0")
        self.grand_label.setStyleSheet(f"font-size:24px;font-weight:800;color:{C['accent']};")
        r3.addWidget(self.grand_label); sb.addLayout(r3)

        ev.addWidget(self.summary_bar)
        self.stack.addWidget(ed)
        rl.addWidget(self.stack)

        splitter.addWidget(right)
        splitter.setSizes([260, 940])
        self.setCentralWidget(splitter)

    def keyPressEvent(self, event: QKeyEvent):
        if event.key() == Qt.Key.Key_Z and event.modifiers() & Qt.KeyboardModifier.ControlModifier:
            self._undo()
            return
        super().keyPressEvent(event)

    # ── 项目列表 ──

    def _load_projects(self):
        self.project_list.blockSignals(True); self.project_list.clear()
        rows = db.fetch("SELECT * FROM projects ORDER BY sort_order ASC, updated_at DESC")
        for r in rows:
            p = dict(r)
            types = [t.strip() for t in p.get("project_types", "print").split(",") if t.strip()]
            if len(types) == 2:
                icon_px = ios_combo_icon(22, C['accent'])
            elif types[0] == "video":
                icon_px = ios_video_icon(22, C['accent'])
            else:
                icon_px = ios_camera_icon(22, C['accent'])
            item = QListWidgetItem(p['name'])
            item.setIcon(QIcon(icon_px))
            item.setData(Qt.ItemDataRole.UserRole, p["id"])
            item.setData(Qt.ItemDataRole.UserRole + 1, p.get("project_types", "print"))
            self.project_list.addItem(item)
        self.project_list.blockSignals(False)

    def _new_project(self):
        dlg = NewProjectDialog(self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            name, types = dlg.get_data()
            pid = db.exec_insert(
                "INSERT INTO projects (name, project_types, currency, tax_rate) VALUES (?,?,'CNY',0)",
                (name, types))
            cats = merge_categories([t.strip() for t in types.split(",")])
            for i, (cn, items) in enumerate(cats):
                cid = db.exec_insert(
                    "INSERT INTO budget_categories (project_id, name, sort_order) VALUES (?,?,?)",
                    (pid, cn, i))
                for j, (desc, note) in enumerate(items):
                    db.exec_insert(
                        "INSERT INTO line_items (category_id, description, notes, sort_order) VALUES (?,?,?,?)",
                        (cid, desc, note, j))
            self._load_projects(); self.project_list.setCurrentRow(0)

    def _delete_project(self):
        idx = self.project_list.currentRow()
        if idx < 0: return
        item = self.project_list.item(idx); pid = item.data(Qt.ItemDataRole.UserRole)
        r = QMessageBox.question(self, "删除", f"删除「{item.text()}」？",
                                 QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if r == QMessageBox.StandardButton.Yes:
            db.exec("DELETE FROM projects WHERE id=?", (pid,))
            if self.current_pid == pid: self._show_empty()
            self._load_projects()

    def _on_projects_reordered(self, parent, start, end, destParent, destRow):
        """拖拽重排后保存新顺序"""
        for i in range(self.project_list.count()):
            item = self.project_list.item(i)
            if item:
                pid = item.data(Qt.ItemDataRole.UserRole)
                db.exec("UPDATE projects SET sort_order=? WHERE id=?", (i, pid))

    def _item_icon(self, item, ptypes, color):
        types = [t.strip() for t in ptypes.split(",") if t.strip()]
        if len(types) == 2:
            item.setIcon(QIcon(ios_combo_icon(22, color)))
        elif types[0] == "video":
            item.setIcon(QIcon(ios_video_icon(22, color)))
        else:
            item.setIcon(QIcon(ios_camera_icon(22, color)))

    def _on_select(self, row):
        if row < 0: return
        # 恢复上一个选中项的图标为黄色
        if hasattr(self, '_last_row') and self._last_row >= 0 and self._last_row < self.project_list.count():
            prev = self.project_list.item(self._last_row)
            if prev:
                self._item_icon(prev, prev.data(Qt.ItemDataRole.UserRole + 1), C['accent'])
        # 当前选中项图标改为黑色（在黄色背景上可见）
        item = self.project_list.item(row)
        self._item_icon(item, item.data(Qt.ItemDataRole.UserRole + 1), "#000000")
        self._last_row = row
        item = self.project_list.item(row)
        self._open_project(item.data(Qt.ItemDataRole.UserRole),
                           item.data(Qt.ItemDataRole.UserRole + 1))

    def _open_project(self, pid, ptypes):
        if pid == self.current_pid: return
        self._timer.stop()
        if self._dirty: self._save()
        self._undo_stack.clear()
        self._update_undo_btn()
        self._loading = True
        self._clear_editor()

        proj = db.fetch_one("SELECT * FROM projects WHERE id=?", (pid,))
        if not proj: return
        p = dict(proj)
        self.current_pid = pid
        self.currency = p.get("currency", "CNY")
        self.tax_rate = Decimal(str(p.get("tax_rate", "0")))

        types = [t.strip() for t in ptypes.split(",") if t.strip()]
        if len(types) == 2:
            self.proj_icon.setPixmap(ios_combo_icon(26, C['accent']))
        elif types[0] == "video":
            self.proj_icon.setPixmap(ios_video_icon(26, C['accent']))
        else:
            self.proj_icon.setPixmap(ios_camera_icon(26, C['accent']))
        self.proj_name_edit.blockSignals(True)
        self.proj_name_edit.setText(p["name"])
        self.proj_name_edit.blockSignals(False)

        idx = self.cur_combo.findData(self.currency)
        if idx >= 0:
            self.cur_combo.blockSignals(True); self.cur_combo.setCurrentIndex(idx)
            self.cur_combo.blockSignals(False)

        ri = next((i for i, (_, r) in enumerate(TAX_RATES) if str(r) == str(self.tax_rate)), 0)
        self.tax_combo.blockSignals(True); self.tax_combo.setCurrentIndex(ri)
        self.tax_combo.blockSignals(False)

        cats = db.fetch(
            "SELECT * FROM budget_categories WHERE project_id=? ORDER BY sort_order", (pid,))
        for cat in cats:
            cd = dict(cat)
            sec = CategorySection(cd, currency=self.currency)
            sec.changed.connect(self._mark_dirty)
            items = db.fetch(
                "SELECT * FROM line_items WHERE category_id=? ORDER BY sort_order", (cd["id"],))
            sec.item_widgets.clear()
            while sec.items_layout.count():
                ch = sec.items_layout.takeAt(0)
                if ch.widget(): ch.widget().deleteLater()
            if items:
                for it in items:
                    li = LineItem(
                        id=it["id"], category_id=it["category_id"],
                        description=it["description"] or "",
                        unit_price=Decimal(str(it["unit_price"] or "0")),
                        quantity=Decimal(str(it["quantity"] or "1")),
                        unit=it["unit"] or "",
                        total=Decimal(str(it["total"] or "0")),
                        notes=it["notes"] or "", sort_order=it["sort_order"] or 0)
                    sec.add_item(li)
            else:
                sec.add_item()
            self.cards_layout.insertWidget(self.cards_layout.count() - 1, sec)
            self.sections.append(sec)

        self.stack.setCurrentIndex(1); self._update_summary()
        self._push_undo()  # 初始快照
        self._loading = False

    def _show_empty(self):
        self._clear_editor(); self.stack.setCurrentIndex(0)

    def _clear_editor(self):
        self.current_pid = None
        for s in self.sections: s.setParent(None); s.deleteLater()
        self.sections.clear()
        while self.cards_layout.count() > 1:
            ch = self.cards_layout.takeAt(0)
            if ch.widget(): ch.widget().deleteLater()

    def _mark_dirty(self):
        if self._loading: return
        self._dirty = True; self._timer.start(500); self._update_summary()

    def _on_name_edit(self):
        if self.current_pid:
            self._dirty = True; self._timer.start(800)

    def _save(self):
        if not self.current_pid or not self._dirty: return
        self._push_db_snapshot()  # 保存前：把 DB 旧状态推入撤销栈
        pid = self.current_pid
        new_name = self.proj_name_edit.text().strip()
        if new_name:
            db.exec("UPDATE projects SET name=?, updated_at=datetime('now','localtime') WHERE id=?",
                    (new_name, pid))
        for i, sec in enumerate(self.sections):
            cd = sec.category_data
            if cd.get("id"):
                db.exec("UPDATE budget_categories SET name=?, sort_order=? WHERE id=?",
                        (cd["name"], i, cd["id"]))
                cid = cd["id"]
            else:
                cid = db.exec_insert(
                    "INSERT INTO budget_categories (project_id, name, sort_order) VALUES (?,?,?)",
                    (pid, cd.get("name", ""), i))
                cd["id"] = cid
            cur = set()
            for j, item in enumerate(sec.get_items()):
                item.category_id = cid; item.sort_order = j
                if item.id:
                    db.exec("""UPDATE line_items SET description=?, unit_price=?, quantity=?, unit=?,
                               total=?, notes=?, sort_order=? WHERE id=?""",
                            (item.description, str(item.unit_price), str(item.quantity),
                             item.unit, str(item.total), item.notes, j, item.id))
                else:
                    item.id = db.exec_insert(
                        """INSERT INTO line_items
                           (category_id, description, unit_price, quantity, unit, total, notes, sort_order)
                           VALUES (?,?,?,?,?,?,?,?)""",
                        (cid, item.description, str(item.unit_price), str(item.quantity),
                         item.unit, str(item.total), item.notes, j))
                cur.add(item.id)
            for row in db.fetch("SELECT id FROM line_items WHERE category_id=?", (cid,)):
                if row["id"] not in cur:
                    db.exec("DELETE FROM line_items WHERE id=?", (row["id"],))
        db.exec("UPDATE projects SET currency=?, tax_rate=?, updated_at=datetime('now','localtime') WHERE id=?",
                (self.currency, str(self.tax_rate), pid))
        self._dirty = False

    def _update_summary(self):
        sub = sum(s.subtotal() for s in self.sections)
        tax = calc_tax(sub, self.tax_rate)
        total = sub + tax
        sym = CURRENCIES.get(self.currency, {}).get("symbol", "¥")
        self.subtotal_label.setText(f"{sym}{sub:,.2f}")
        self.tax_label.setText(f"{sym}{tax:,.2f}")
        self.grand_label.setText(f"{sym}{total:,.2f}")

    # ── 撤销 ──

    def _push_db_snapshot(self):
        """从数据库读取当前状态推入撤销栈（保存前用）"""
        if self._undoing or not self.current_pid:
            return
        snap = {"currency": self.currency, "tax_rate": str(self.tax_rate), "categories": []}
        cats = db.fetch("SELECT * FROM budget_categories WHERE project_id=? ORDER BY sort_order", (self.current_pid,))
        for cat in cats:
            cd = dict(cat)
            items = db.fetch("SELECT * FROM line_items WHERE category_id=? ORDER BY sort_order", (cd["id"],))
            cat_snap = {"id": cd["id"], "name": cd["name"], "sort_order": cd["sort_order"], "items": []}
            for it in items:
                cat_snap["items"].append({
                    "description": it["description"] or "",
                    "unit_price": str(it["unit_price"] or "0"),
                    "quantity": str(it["quantity"] or "1"),
                    "unit": it["unit"] or "",
                    "total": str(it["total"] or "0"),
                    "sort_order": it["sort_order"] or 0,
                })
            snap["categories"].append(cat_snap)
        self._undo_stack.append(snap)
        self._update_undo_btn()

    def _push_undo(self):
        """保存当前 UI 状态快照到撤销栈"""
        if self._undoing or not self.current_pid:
            return
        snap = {
            "currency": self.currency,
            "tax_rate": str(self.tax_rate),
            "categories": [],
        }
        for sec in self.sections:
            cat_snap = {
                "id": sec.category_data.get("id"),
                "name": sec.category_data.get("name", ""),
                "sort_order": sec.category_data.get("sort_order", 0),
                "items": [it.get_data().model_dump() for it in sec.item_widgets],
            }
            # Convert Decimal to str for JSON compatibility
            for it in cat_snap["items"]:
                for k in ("unit_price", "quantity", "total"):
                    it[k] = str(it.get(k, "0"))
            snap["categories"].append(cat_snap)
        self._undo_stack.append(snap)
        self._update_undo_btn()

    def _undo(self):
        """Ctrl+Z / 按钮 恢复上一个快照（保留最后一个快照不清空）"""
        if len(self._undo_stack) <= 1 or not self.current_pid:
            return
        self._undoing = True
        snap = self._undo_stack.pop()
        self.statusBar().showMessage(f"已撤销，剩余 {len(self._undo_stack)} 步", 2000)
        self._timer.stop()

        # 恢复币种和税率
        self.currency = snap["currency"]
        self.tax_rate = Decimal(snap["tax_rate"])
        idx = self.cur_combo.findData(self.currency)
        if idx >= 0:
            self.cur_combo.blockSignals(True); self.cur_combo.setCurrentIndex(idx)
            self.cur_combo.blockSignals(False)
        ri = next((i for i, (_, r) in enumerate(TAX_RATES) if str(r) == str(self.tax_rate)), 0)
        self.tax_combo.blockSignals(True); self.tax_combo.setCurrentIndex(ri)
        self.tax_combo.blockSignals(False)

        # 清除当前 sections
        for s in self.sections:
            s.setParent(None); s.deleteLater()
        self.sections.clear()
        while self.cards_layout.count() > 1:
            ch = self.cards_layout.takeAt(0)
            if ch.widget(): ch.widget().deleteLater()

        # 从快照重建
        for cat_snap in snap["categories"]:
            cd = {"id": cat_snap["id"], "project_id": self.current_pid,
                  "name": cat_snap["name"], "sort_order": cat_snap["sort_order"]}
            sec = CategorySection(cd, currency=self.currency)
            sec.changed.connect(self._mark_dirty)
            sec.item_widgets.clear()
            while sec.items_layout.count():
                ch = sec.items_layout.takeAt(0)
                if ch.widget(): ch.widget().deleteLater()
            for it_data in cat_snap["items"]:
                li = LineItem(
                    description=it_data.get("description", ""),
                    unit_price=Decimal(it_data.get("unit_price", "0")),
                    quantity=Decimal(it_data.get("quantity", "1")),
                    unit=it_data.get("unit", ""),
                    total=Decimal(it_data.get("total", "0")),
                    sort_order=it_data.get("sort_order", 0))
                sec.add_item(li)
            if not sec.item_widgets:
                sec.add_item()
            self.cards_layout.insertWidget(self.cards_layout.count() - 1, sec)
            self.sections.append(sec)

        self._update_summary()
        self._save()
        self._undoing = False
        self._update_undo_btn()

    def _update_undo_btn(self):
        n = len(self._undo_stack)
        self.undo_btn.setEnabled(n > 1)
        self.undo_btn.setText(f"↩ 撤销 ({n - 1})" if n > 1 else "↩ 撤销")

    def _on_currency_change(self):
        code = self.cur_combo.currentData()
        if code and code != self.currency:
            self.currency = code
            for s in self.sections: s.set_currency(code)
            self._update_summary(); self._mark_dirty()

    def _on_tax_change(self):
        rs = self.tax_combo.currentData()
        if rs:
            self.tax_rate = Decimal(rs)
            self._update_summary(); self._mark_dirty()

    def _get_proj(self):
        r = db.fetch_one("SELECT * FROM projects WHERE id=?", (self.current_pid,))
        return dict(r) if r else None

    # ── 粘贴导入 ──

    def _import_paste(self):
        if not self.current_pid: return
        dlg = QDialog(self)
        dlg.setWindowTitle("粘贴导入")
        dlg.setFixedSize(600, 450)
        dlg.setModal(True)
        dlg.setStyleSheet(f"""
            QDialog {{ background: #1C1C1E; border-radius: 14px; }}
            QLabel {{ color: {C['text']}; }}
            QTextEdit {{ background: #000000; color: {C['text']}; border: 1px solid {C['separator']};
                        border-radius: 8px; padding: 12px; font-size: 14px;
                        selection-background-color: {C['accent']}; selection-color: #000000; }}
        """)
        lo = QVBoxLayout(dlg)
        lo.setContentsMargins(20, 20, 20, 16); lo.setSpacing(12)
        t = QLabel("粘贴费用明细（从 PDF / Excel / 微信 复制后在此粘贴）")
        t.setStyleSheet(f"font-size:14px;font-weight:600;color:{C['secondary']};")
        lo.addWidget(t)
        edit = QTextEdit()
        edit.setPlaceholderText("在此粘贴...")
        lo.addWidget(edit, 1)
        btn_row = QHBoxLayout(); btn_row.addStretch()
        cancel = QPushButton("取消")
        cancel.setStyleSheet(f"background:transparent;color:{C['secondary']};border:none;font-size:14px;font-weight:600;padding:10px 18px;")
        cancel.clicked.connect(dlg.reject); btn_row.addWidget(cancel)
        confirm = QPushButton("解析并预览")
        confirm.setStyleSheet(f"""
            QPushButton {{ background: {C['accent']}; color: #000000; border: none;
                          border-radius: 20px; padding: 10px 22px; font-size: 14px; font-weight: 700; }}
            QPushButton:hover {{ background: {C['accent_hover']}; }}
            QPushButton:pressed {{ background: #E5C008; }}
        """)
        confirm.clicked.connect(dlg.accept); btn_row.addWidget(confirm)
        lo.addLayout(btn_row)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        text = edit.toPlainText().strip()
        if not text:
            QMessageBox.information(self, "提示", "未输入任何内容。")
            return
        items = parse_clipboard_text(text)
        if not items:
            QMessageBox.information(self, "解析结果",
                f"未能识别到费用明细。\n\n输入内容预览:\n{text[:300]}")
            return
        if not self._show_import_preview(items):
            return
        self._populate_from_items(items)

    # ── Excel 导入 ──

    def _import_excel(self):
        if not self.current_pid: return
        path, _ = QFileDialog.getOpenFileName(
            self, "导入 Excel 报价单", "", "Excel 文件 (*.xlsx *.xls)")
        if not path: return
        try:
            items = import_excel(path)
        except Exception as e:
            QMessageBox.critical(self, "导入失败", str(e))
            log.exception("Excel import failed")
            return
        if not items:
            QMessageBox.information(self, "导入结果", "未能从 Excel 识别到费用明细。")
            return
        if not self._show_import_preview(items):
            return
        self._populate_from_items(items)

    # ── 共用：预览 + 填入 ──

    def _show_import_preview(self, items: list) -> bool:
        cat_count = len(set(it.get("category", "") for it in items))
        grand_total = sum(it["total"] for it in items)

        dlg = QDialog(self)
        dlg.setWindowTitle("导入预览")
        dlg.setFixedSize(640, 500)
        dlg.setModal(True)
        dlg.setStyleSheet(f"""
            QDialog {{ background: #1C1C1E; border-radius: 14px; }}
            QLabel {{ color: {C['text']}; }}
        """)

        lo = QVBoxLayout(dlg)
        lo.setContentsMargins(20, 20, 20, 16); lo.setSpacing(12)

        t = QLabel("已识别以下费用明细：")
        t.setStyleSheet(f"font-size:16px;font-weight:700;color:{C['text']};")
        lo.addWidget(t)

        info = QLabel(f"共 {len(items)} 条费用 | {cat_count} 个分类 | 总计: ¥{grand_total:,.2f}")
        info.setStyleSheet(f"font-size:12px;color:{C['secondary']};font-weight:600;")
        lo.addWidget(info)

        table = QTableWidget()
        table.setColumnCount(4)
        table.setHorizontalHeaderLabels(["描述", "单价", "数量", "合计"])
        table.setRowCount(min(len(items), 50))
        table.setEditTriggers(table.EditTrigger.NoEditTriggers)
        table.setSelectionBehavior(table.SelectionBehavior.SelectRows)
        table.horizontalHeader().setStretchLastSection(True)
        table.setStyleSheet(f"""
            QTableWidget {{ background: #000000; border: 1px solid {C['separator']};
                           border-radius: 8px; gridline-color: {C['separator']}; color: {C['text']}; }}
            QHeaderView::section {{ background: #2C2C2E; color: {C['secondary']};
                                   font-weight: 700; padding: 6px; border: none; }}
        """)

        sym = CURRENCIES.get(self.currency, {}).get("symbol", "¥")
        for i, it in enumerate(items[:50]):
            table.setItem(i, 0, QTableWidgetItem(it["description"]))
            table.setItem(i, 1, QTableWidgetItem(f"{sym}{it['unit_price']:,.2f}"))
            qty_str = f"{it['quantity']:g}{it.get('unit','')}"
            table.setItem(i, 2, QTableWidgetItem(qty_str))
            table.setItem(i, 3, QTableWidgetItem(f"{sym}{it['total']:,.2f}"))

        for r in range(table.rowCount()):
            table.setRowHeight(r, 28)
        table.setColumnWidth(0, 280)
        table.setColumnWidth(1, 100)
        table.setColumnWidth(2, 100)

        lo.addWidget(table, 1)

        btn_row = QHBoxLayout(); btn_row.addStretch()
        cancel = QPushButton("取消")
        cancel.setStyleSheet(f"background:transparent;color:{C['secondary']};border:none;font-size:14px;font-weight:600;padding:10px 18px;")
        cancel.clicked.connect(dlg.reject); btn_row.addWidget(cancel)
        confirm = QPushButton("确认导入")
        confirm.setStyleSheet(f"""
            QPushButton {{ background: {C['accent']}; color: #000000; border: none;
                          border-radius: 20px; padding: 10px 22px; font-size: 14px; font-weight: 700; }}
            QPushButton:hover {{ background: {C['accent_hover']}; }}
            QPushButton:pressed {{ background: #E5C008; }}
        """)
        confirm.clicked.connect(dlg.accept); btn_row.addWidget(confirm)
        lo.addLayout(btn_row)

        return dlg.exec() == QDialog.DialogCode.Accepted

    def _populate_from_items(self, items: list):
        self._push_undo()
        self._timer.stop()
        if self._dirty: self._save()

        cat_map = {}
        for item in items:
            cat = item.get("category", "导入费用")
            if cat not in cat_map: cat_map[cat] = []
            cat_map[cat].append(item)

        for s in self.sections:
            s.setParent(None); s.deleteLater()
        self.sections.clear()
        while self.cards_layout.count() > 1:
            ch = self.cards_layout.takeAt(0)
            if ch.widget(): ch.widget().deleteLater()

        for i, (cat_name, cat_items) in enumerate(cat_map.items()):
            existing = db.fetch_one(
                "SELECT id FROM budget_categories WHERE project_id=? AND name=?",
                (self.current_pid, cat_name))
            if existing:
                cid = existing["id"]
            else:
                cid = db.exec_insert(
                    "INSERT INTO budget_categories (project_id, name, sort_order) VALUES (?,?,?)",
                    (self.current_pid, cat_name, i))
            sec = CategorySection({"id": cid, "name": cat_name, "project_id": self.current_pid, "sort_order": i},
                                  currency=self.currency)
            sec.changed.connect(self._mark_dirty)
            sec.item_widgets.clear()
            while sec.items_layout.count():
                ch = sec.items_layout.takeAt(0)
                if ch.widget(): ch.widget().deleteLater()
            for j, it in enumerate(cat_items):
                li = LineItem(
                    description=it["description"],
                    unit_price=m(it["unit_price"]),
                    quantity=m(it["quantity"]),
                    unit=it.get("unit", ""),
                    total=m(it["total"]),
                    sort_order=j)
                sec.add_item(li)
            if not sec.item_widgets:
                sec.add_item()
            self.cards_layout.insertWidget(self.cards_layout.count() - 1, sec)
            self.sections.append(sec)

        self._update_summary()
        self._mark_dirty()
        self.stack.setCurrentIndex(1)

    # ── 导出 Excel ──
    # ── 导出 Excel ──

    def _export_excel(self):
        if not self.current_pid: return
        self._timer.stop(); self._save()
        proj = self._get_proj()
        if not proj: return
        sym = CURRENCIES.get(proj.get("currency", "CNY"), {}).get("symbol", "¥")
        path, _ = QFileDialog.getSaveFileName(
            self, "导出 Excel", f"费用_{proj['name']}_{datetime.now():%Y%m%d}.xlsx", "Excel (*.xlsx)")
        if not path: return
        try:
            self._do_xlsx(path, proj, sym)
            QMessageBox.information(self, "导出成功", f"已导出:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "导出失败", str(e))
            log.exception("Excel export")

    def _do_xlsx(self, path, proj, sym):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook(); ws = wb.active; ws.title = "费用明细"
        black = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        dark = PatternFill(start_color="1C1C1E", end_color="1C1C1E", fill_type="solid")
        cat_fill = PatternFill(start_color="2C2C2E", end_color="2C2C2E", fill_type="solid")
        thin = Border(bottom=Side(style="hair", color="38383A"))

        ws.merge_cells("A1:G1")
        ws["A1"] = f"{proj['name']}  费用明细"
        ws["A1"].font = Font(name="Microsoft YaHei", size=18, bold=True, color="FFFFFF")
        ws["A1"].alignment = Alignment(horizontal="center")
        ws["A1"].fill = black; ws.row_dimensions[1].height = 36

        ws.merge_cells("A2:G2")
        types = [t.strip() for t in proj.get("project_types", "print").split(",")]
        tl = "套拍 (平面+视频)" if len(types) == 2 else ("平面" if types[0] == "print" else "视频")
        ws["A2"] = f"{tl} | 币种:{proj.get('currency','CNY')} | {datetime.now():%Y-%m-%d %H:%M}"
        ws["A2"].font = Font(color="8E8E93", size=10)
        ws["A2"].alignment = Alignment(horizontal="center"); ws["A2"].fill = black

        for c, h in enumerate(["分类", "费用项", "单价", "数量", "单位", "合计", "备注"], 1):
            cell = ws.cell(row=4, column=c, value=h)
            cell.font = Font(name="Microsoft YaHei", size=10, bold=True, color="98989D")
            cell.fill = dark; cell.alignment = Alignment(horizontal="center")

        row = 5; gs = Decimal("0")
        cats = db.fetch("SELECT * FROM budget_categories WHERE project_id=? ORDER BY sort_order",
                        (self.current_pid,))
        for cat in cats:
            cd = dict(cat)
            items = db.fetch("SELECT * FROM line_items WHERE category_id=? ORDER BY sort_order",
                             (cd["id"],))
            ct = sum(Decimal(str(it["total"] or "0")) for it in items); gs += ct
            ws.merge_cells(f"A{row}:G{row}")
            ws.cell(row=row, column=1, value=cd["name"]).font = Font(
                name="Microsoft YaHei", size=12, bold=True, color="FFFFFF")
            for c in range(1, 8): ws.cell(row=row, column=c).fill = cat_fill
            ws.cell(row=row, column=6, value=f"小计: {sym}{ct:,.2f}").font = Font(
                bold=True, color="98989D", size=11)
            ws.cell(row=row, column=6).fill = cat_fill
            ws.cell(row=row, column=6).alignment = Alignment(horizontal="right")
            row += 1
            for it in items:
                d = dict(it); t = Decimal(str(d["total"] or "0"))
                ws.cell(row=row, column=2, value=d["description"] or "").font = Font(color="FFFFFF", size=11)
                ws.cell(row=row, column=3, value=float(Decimal(str(d["unit_price"] or "0"))))
                ws.cell(row=row, column=3).number_format = f'{sym}#,##0.00'
                ws.cell(row=row, column=4, value=float(Decimal(str(d["quantity"] or "1"))))
                ws.cell(row=row, column=5, value=d["unit"] or "")
                ws.cell(row=row, column=6, value=float(t))
                ws.cell(row=row, column=6).number_format = f'{sym}#,##0.00'
                ws.cell(row=row, column=7, value=d["notes"] or "")
                for c in range(1, 8): ws.cell(row=row, column=c).border = thin
                row += 1
            row += 1

        row += 1; tr = Decimal(str(proj.get("tax_rate", "0")))
        ta = calc_tax(gs, tr); gt = gs + ta
        for lb, am, fn in [
            ("不含税合计:", gs, Font(name="Microsoft YaHei", size=12, bold=True, color="FFFFFF")),
            (f"税额 ({float(tr)*100:.0f}%):", ta, Font(name="Microsoft YaHei", size=11, color="FF9F0A")),
            ("含税总计:", gt, Font(name="Microsoft YaHei", size=14, bold=True, color="FFD60A")),
        ]:
            ws.merge_cells(f"A{row}:D{row}")
            ws.cell(row=row, column=1, value=lb).font = fn
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="right")
            ws.cell(row=row, column=6, value=float(am)).font = fn
            ws.cell(row=row, column=6).number_format = f'{sym}#,##0.00'
            ws.cell(row=row, column=6).alignment = Alignment(horizontal="right")
            row += 1

        for i, w in enumerate([18, 32, 14, 8, 8, 16, 20], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        wb.save(path)

    # ── 导出 PDF ──

    def _export_pdf(self):
        if not self.current_pid: return
        self._timer.stop(); self._save()
        proj = self._get_proj()
        if not proj: return
        sym = CURRENCIES.get(proj.get("currency", "CNY"), {}).get("symbol", "¥")
        path, _ = QFileDialog.getSaveFileName(
            self, "导出 PDF", f"费用_{proj['name']}_{datetime.now():%Y%m%d}.pdf", "PDF (*.pdf)")
        if not path: return
        try:
            self._do_pdf(path, proj, sym)
            QMessageBox.information(self, "导出成功", f"已导出:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "导出失败", str(e))
            log.exception("PDF export")

    def _do_pdf(self, path, proj, sym):
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.lib.colors import HexColor
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable)
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont

        cjk = "Helvetica"
        for fp in ["C:/Windows/Fonts/msyh.ttc", "C:/Windows/Fonts/msyhbd.ttc",
                    "C:/Windows/Fonts/simhei.ttf"]:
            if Path(fp).exists():
                try:
                    pdfmetrics.registerFont(TTFont("CJK", fp, subfontIndex=0))
                    cjk = "CJK"; break
                except Exception: continue

        doc = SimpleDocTemplate(path, pagesize=A4,
                                leftMargin=20*mm, rightMargin=20*mm,
                                topMargin=20*mm, bottomMargin=20*mm)
        story = []

        ts = ParagraphStyle("TS", fontName=cjk, fontSize=20, leading=28,
                            spaceAfter=2*mm, alignment=TA_CENTER, textColor=HexColor("#FFFFFF"))
        ss = ParagraphStyle("SS", fontName=cjk, fontSize=10, leading=14,
                            spaceAfter=10*mm, alignment=TA_CENTER, textColor=HexColor("#98989D"))
        cs = ParagraphStyle("CS", fontName=cjk, fontSize=11, leading=16,
                            textColor=HexColor("#98989D"), spaceBefore=4*mm, spaceAfter=2*mm)
        its = ParagraphStyle("IS", fontName=cjk, fontSize=10, leading=14, textColor=HexColor("#FFFFFF"))
        ns = ParagraphStyle("NS", fontName=cjk, fontSize=8, leading=12, textColor=HexColor("#636366"))
        fs = ParagraphStyle("FS", fontName=cjk, fontSize=8, leading=12,
                            textColor=HexColor("#48484A"), alignment=TA_CENTER)

        types = [t.strip() for t in proj.get("project_types", "print").split(",")]
        tl_text = "套拍 (平面 + 视频)" if len(types) == 2 else ("平面" if types[0] == "print" else "视频")

        story.append(Paragraph("项目费用结算", ts))
        story.append(Paragraph(proj['name'], ts))
        story.append(Paragraph(
            f"{tl_text} | 币种: {proj.get('currency','CNY')} {sym} | {datetime.now():%Y-%m-%d}", ss))
        story.append(HRFlowable(width="100%", thickness=0.5, color=HexColor("#38383A")))
        story.append(Spacer(1, 4*mm))

        cats = db.fetch("SELECT * FROM budget_categories WHERE project_id=? ORDER BY sort_order",
                        (self.current_pid,))
        gs = Decimal("0"); data = []
        for cat in cats:
            cd = dict(cat)
            items = db.fetch("SELECT * FROM line_items WHERE category_id=? ORDER BY sort_order",
                             (cd["id"],))
            ct = sum(Decimal(str(it["total"] or "0")) for it in items); gs += ct
            data.append([
                Paragraph(f"<b>{cd['name']}</b>", cs), "", "", "",
                Paragraph(f"<b>{sym}{ct:,.2f}</b>",
                          ParagraphStyle("R", fontName=cjk, fontSize=10,
                                         textColor=HexColor("#98989D"), alignment=TA_RIGHT)),
                ""])
            for it in items:
                d = dict(it); t = Decimal(str(d["total"] or "0"))
                u = (d.get("unit") or "").strip()
                desc = (d["description"] or "").strip() or "-"
                note = (d["notes"] or "").strip()
                data.append([
                    Paragraph(desc, its),
                    Paragraph(f"{sym}{float(Decimal(str(d['unit_price'] or '0'))):,.2f}", its),
                    Paragraph(f"{float(Decimal(str(d['quantity'] or '1'))):g}{u}", its),
                    "",
                    Paragraph(f"{sym}{float(t):,.2f}", its),
                    Paragraph(note, ns) if note else "",
                ])

        cw = [72*mm, 28*mm, 24*mm, 4*mm, 28*mm, 24*mm]
        tbl = Table(data, colWidths=cw, repeatRows=0)
        tbl.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('LINEBELOW', (0, 0), (-1, -1), 0.3, HexColor("#2C2C2E")),
        ]))
        story.append(tbl)
        story.append(Spacer(1, 8*mm))
        story.append(HRFlowable(width="100%", thickness=0.5, color=HexColor("#FFD60A")))
        story.append(Spacer(1, 4*mm))

        tr = Decimal(str(proj.get("tax_rate", "0")))
        ta = calc_tax(gs, tr); gt = gs + ta
        sd = [
            ["不含税合计:", f"{sym}{gs:,.2f}"],
            [f"税额 ({float(tr)*100:.0f}%):", f"{sym}{ta:,.2f}"],
            ["含税总计:", f"{sym}{gt:,.2f}"],
        ]
        st = Table(sd, colWidths=[140*mm, 40*mm])
        st.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('FONTNAME', (0, 0), (-1, -1), cjk),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('TEXTCOLOR', (0, 0), (-1, -1), HexColor("#FFFFFF")),
            ('TEXTCOLOR', (1, 1), (1, 1), HexColor("#FF9F0A")),
            ('TEXTCOLOR', (1, 2), (1, 2), HexColor("#FFD60A")),
            ('FONTSIZE', (0, 2), (1, 2), 14),
        ]))
        story.append(st)
        story.append(Spacer(1, 12*mm))
        story.append(HRFlowable(width="100%", thickness=0.5, color=HexColor("#38383A")))
        story.append(Paragraph(
            f"系统自动生成，仅供内部对账使用 | {datetime.now():%Y-%m-%d %H:%M}", fs))

        doc.build(story)


def main():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    app.setPalette(DARK_PALETTE)
    app.setFont(QFont("Microsoft YaHei", 10))
    app.setStyleSheet(STYLE)
    w = App(); w._show_empty(); w.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
